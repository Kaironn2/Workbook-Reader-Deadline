import openpyxl

# Carregar a planilha original
workbook = openpyxl.load_workbook('C:\\Users\\jonat\\OneDrive\\Área de Trabalho\\Trabalho\\planilhas\\Prazo PME 2024-05.xlsx')
sheet = workbook.active

dados = {} # dicionário p/ armazenar estado

# Iterar sobre as linhas da planilha
for row in sheet.iter_rows(min_row=2, values_only=True):  # Começar da linha 2 para ignorar o cabeçalho
    municipio, estado, prazo = row[2], row[3], row[24]  # Ajustar índices das colunas (C=2, D=3, Y=24)

    if estado not in dados: # verifica se o estado já está em dados, se n estiver, adiciona
        dados[estado] = {}

    if municipio not in dados[estado]: # verifica se o municipio já está em estado
        dados[estado][municipio] = prazo # se estiver, adiciona o prazo atual
    else:
        # Atualiza o prazo se o novo prazo for maior
        if prazo > dados[estado][municipio]:
            dados[estado][municipio] = prazo

# Criar novo workbook pra exportar os dados
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Novo prazo"

# Criar cabeçalho
new_sheet.append(["Estado", "Cidade", "Prazo"])

# Preencher a nova planilha com os dados tratados
for estado, cidades in dados.items():
    for cidade, prazo in cidades.items():
        new_sheet.append([estado, cidade, prazo])

# Salvar arquivo
new_workbook.save('C:\\Users\\jonat\\OneDrive\\Área de Trabalho\\Trabalho\\planilhas\\Prazo_Atual.xlsx')
print("Dados exportados com sucesso!")
